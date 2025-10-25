from flask import Blueprint, jsonify, request, send_file
from firebase_utils import get_firestore_client
from firebase_admin import firestore
from datetime import datetime
import traceback

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib import colors
from reportlab.lib.units import inch
from statistics import mean
from openpyxl import Workbook, cell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference

projects_bp = Blueprint('projects', __name__)

# =============== HELPER FUNCTION TO CHECK PROJECT COMPLETION ===============
def is_project_completed(project_id, db):
    """
    Check if all tasks and subtasks in a project are completed.
    Returns True if project should be hidden, False otherwise.
    """
    try:
        # Get all tasks for this project
        tasks_ref = db.collection('Tasks')
        tasks_query = tasks_ref.where('proj_ID', '==', project_id)
        tasks = list(tasks_query.stream())
        
        # If no tasks exist, project is not complete
        if not tasks:
            return False
        
        # Check if all tasks are completed
        for task_doc in tasks:
            task_data = task_doc.to_dict()
            
            # Skip deleted tasks
            if task_data.get('is_deleted', False):
                continue
            
            # If any task is not completed, project is not complete
            if task_data.get('task_status') != 'Completed':
                return False
            
            # Check subtasks if they exist
            subtasks = task_data.get('subtasks', [])
            for subtask in subtasks:
                # Skip deleted subtasks
                if subtask.get('is_deleted', False):
                    continue
                    
                # If any subtask is not completed, project is not complete
                if subtask.get('status') != 'Completed':
                    return False
        
        return True
    except Exception as e:
        print(f"Error checking project completion: {str(e)}")
        return False

# =============== CALENDAR GENERATION HELPER FUNCTIONS ===============
def generate_team_calendar_pdf(tasks, project_name, project_data, users_ref):
    """Generate team schedule PDF with tasks for each collaborator"""
    users = {u.id: u.to_dict() for u in users_ref.stream()}
    styles = getSampleStyleSheet()
    
    # Get project collaborators
    project_collaborator_ids = set(project_data.get('collaborators', []))
    
    # Build task list for each collaborator
    collaborator_tasks = {}
    
    for task in tasks:
        assigned_to = task.get("assigned_to", [])
        task_name = task.get("task_name", "Unknown Task")
        priority = task.get("priority_level", 1)
        
        # Parse dates
        start_date = task.get("start_date", "")
        end_date = task.get("end_date", "")
        
        duration = "N/A"
        start_str = "N/A"
        end_str = "N/A"
        
        if start_date:
            try:
                if isinstance(start_date, str):
                    if "T" in start_date:
                        dt = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                    else:
                        dt = datetime.fromisoformat(start_date)
                else:
                    dt = start_date
                start_str = dt.strftime("%Y-%m-%d")
            except:
                pass
        
        if end_date:
            try:
                if isinstance(end_date, str):
                    if "T" in end_date:
                        dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                    else:
                        dt = datetime.fromisoformat(end_date)
                else:
                    dt = end_date
                end_str = dt.strftime("%Y-%m-%d")
                
                # Calculate duration
                if start_date and end_date:
                    try:
                        start_dt = datetime.fromisoformat(start_date.replace("Z", "+00:00")) if isinstance(start_date, str) else start_date
                        end_dt = datetime.fromisoformat(end_date.replace("Z", "+00:00")) if isinstance(end_date, str) else end_date
                        if hasattr(start_dt, 'date'):
                            start_dt = start_dt.date()
                        if hasattr(end_dt, 'date'):
                            end_dt = end_dt.date()
                        delta = end_dt - start_dt
                        duration = f"{delta.days + 1} days"
                    except:
                        duration = "N/A"
            except:
                pass
        
        # Assign task to each collaborator
        for user_id in assigned_to:
            if user_id in project_collaborator_ids:
                user_name = users.get(user_id, {}).get('name', f'User_{user_id}')
                
                if user_name not in collaborator_tasks:
                    collaborator_tasks[user_name] = []
                
                collaborator_tasks[user_name].append({
                    'name': task_name,
                    'start': start_str,
                    'end': end_str,
                    'duration': duration,
                    'priority': priority
                })
    
    # Generate PDF elements
    calendar_elements = []
    
    # Add all collaborators (even if they have no tasks)
    all_collaborators = {}
    for user_id in project_collaborator_ids:
        user_name = users.get(user_id, {}).get('name', f'User_{user_id}')
        all_collaborators[user_name] = collaborator_tasks.get(user_name, [])
    
    # Sort collaborators alphabetically
    sorted_collaborators = sorted(all_collaborators.keys())
    
    # Create table for each collaborator
    for collaborator_name in sorted_collaborators:
        tasks_list = all_collaborators[collaborator_name]
        
        # Collaborator header
        collaborator_header = Paragraph(f"<b>{collaborator_name}</b>", styles["h2"])
        calendar_elements.append(collaborator_header)
        calendar_elements.append(Spacer(1, 8))
        
        if not tasks_list:
            no_tasks_text = Paragraph("<i>No tasks assigned</i>", styles["Normal"])
            calendar_elements.append(no_tasks_text)
            calendar_elements.append(Spacer(1, 16))
            continue
        
        # Create table with task details
        table_data = [["Task Name", "Start Date", "End Date", "Duration", "Priority"]]
        
        for task in tasks_list:
            priority_str = "High" if task['priority'] >= 4 else "Medium" if task['priority'] >= 2 else "Low"
            table_data.append([
                task['name'],
                task['start'],
                task['end'],
                task['duration'],
                priority_str
            ])
        
        task_table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1*inch, 0.8*inch, 0.7*inch])
        task_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.beige, colors.white]),
        ]))
        
        calendar_elements.append(task_table)
        calendar_elements.append(Spacer(1, 20))
    
    return calendar_elements


def generate_calendar_months(tasks, project_name):
    """Generate calendar months with tasks for PDF"""
    from collections import defaultdict
    import calendar
    
    # Group tasks by month
    tasks_by_month = defaultdict(list)
    
    for task in tasks:
        end_date = task.get("end_date")
        if end_date:
            try:
                if isinstance(end_date, str):
                    if "T" in end_date:
                        dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                    else:
                        dt = datetime.fromisoformat(end_date)
                else:
                    dt = end_date
                
                month_key = (dt.year, dt.month)
                tasks_by_month[month_key].append(task)
            except Exception:
                continue
    
    calendar_elements = []
    
    # Generate calendar for each month with tasks
    for (year, month), month_tasks in sorted(tasks_by_month.items()):
        # Month header
        month_name = calendar.month_name[month]
        month_header = Paragraph(f"<b>{month_name} {year}</b>", getSampleStyleSheet()["h2"])
        
        # Create calendar grid
        cal = calendar.monthcalendar(year, month)
        calendar_data = []
        
        # Header row
        header_row = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        calendar_data.append(header_row)
        
        # Calendar rows
        for week in cal:
            week_row = []
            for day in week:
                if day == 0:
                    week_row.append("")
                else:
                    # Check if there are tasks on this day
                    day_tasks = []
                    for task in month_tasks:
                        try:
                            task_date = task.get("end_date")
                            if isinstance(task_date, str):
                                if "T" in task_date:
                                    dt = datetime.fromisoformat(task_date.replace("Z", "+00:00"))
                                else:
                                    dt = datetime.fromisoformat(task_date)
                            else:
                                dt = task_date
                            
                            if dt.day == day:
                                # Color code by priority
                                priority = task.get("priority_level", 1)
                                if priority >= 4:
                                    color = "red"
                                elif priority >= 3:
                                    color = "orange" 
                                else:
                                    color = "green"
                                
                                task_name = task.get("task_name", "Unknown")[:15]  # Truncate long names
                                day_tasks.append(f'<font color="{color}">• {task_name}</font>')
                        except Exception:
                            continue
                    
                    if day_tasks:
                        day_content = f"{day}<br/>{'<br/>'.join(day_tasks)}"
                    else:
                        day_content = str(day)
                    
                    week_row.append(day_content)
            
            calendar_data.append(week_row)
        
        # Create calendar table
        calendar_table = Table(calendar_data, colWidths=[1.2*inch]*7, rowHeights=[0.4*inch]*len(calendar_data))
        calendar_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
        ]))
        
        calendar_elements.append(KeepTogether([month_header, Spacer(1, 12), calendar_table]))
    
    return calendar_elements

# =============== HELPER FUNCTION FOR DEPARTMENT FILTERING ===============
def get_users_from_same_division(division_name):
    """Get all user IDs from the same division"""
    try:
        db = get_firestore_client()
        users_ref = db.collection('Users')
        users_query = users_ref.where('division_name', '==', division_name)
        users = users_query.stream()
        
        user_ids = []
        for user in users:
            user_ids.append(user.id)
        
        print(f"Found {len(user_ids)} users in {division_name} division")
        return user_ids
    except Exception as e:
        print(f"Error getting users from division {division_name}: {str(e)}")
        return []

@projects_bp.route('/api/projects/filtered/<division_name>', methods=['GET'])
def get_filtered_projects_by_division(division_name):
    """Get projects filtered by division - only shows projects where current user is a collaborator"""
    try:
        db = get_firestore_client()
        
        # Get current user ID from query parameters
        current_user_id = request.args.get('user_id')
        if not current_user_id:
            print("No user ID provided in query parameters")
            return jsonify({'error': 'User ID required'}), 400
        
        # Get show_completed parameter (default to False to hide completed projects)
        show_completed = request.args.get('show_completed', 'false').lower() == 'true'
        
        print(f"Filtering projects for division: {division_name}, user: {current_user_id}, show_completed: {show_completed}")
        
        # Get all projects
        projects_ref = db.collection('Projects')
        all_projects = projects_ref.stream()
        
        filtered_projects = []
        
        for project in all_projects:
            project_data = project.to_dict()
            project_data['id'] = project.id
            
            # Check if current user is a collaborator of this project
            collaborators = project_data.get('collaborators', [])
            is_user_collaborator = current_user_id in collaborators
            
            if is_user_collaborator:
                project_id = project.id
                
                # Check if project is completed
                is_complete = is_project_completed(project_id, db)
                
                # Skip completed projects unless show_completed is True
                if is_complete and not show_completed:
                    print(f"Project {project_data.get('proj_name', 'Unknown')} excluded - completed and filter is off")
                    continue
                
                print(f"Project {project_data.get('proj_name', 'Unknown')} included - user is collaborator")
                
                # Convert timestamps to ISO format
                if 'start_date' in project_data and project_data['start_date']:
                    project_data['start_date'] = project_data['start_date'].isoformat()
                if 'end_date' in project_data and project_data['end_date']:
                    project_data['end_date'] = project_data['end_date'].isoformat()
                if 'createdAt' in project_data and project_data['createdAt']:
                    project_data['createdAt'] = project_data['createdAt'].isoformat()
                if 'updatedAt' in project_data and project_data['updatedAt']:
                    project_data['updatedAt'] = project_data['updatedAt'].isoformat()
                
                # Get all tasks for this project (since user is a collaborator)
                project_doc_id = project.id
                tasks_ref = db.collection('Tasks')
                tasks_query = tasks_ref.where('proj_ID', '==', project_doc_id)
                tasks = tasks_query.stream()
                
                task_list = []
                for task in tasks:
                    task_data = task.to_dict()
                    task_data['id'] = task.id

                    if task_data.get('is_deleted', False):
                        continue
                    
                    # Convert task timestamps
                    if 'start_date' in task_data and task_data['start_date']:
                        task_data['start_date'] = task_data['start_date'].isoformat()
                    if 'end_date' in task_data and task_data['end_date']:
                        task_data['end_date'] = task_data['end_date'].isoformat()
                    if 'createdAt' in task_data and task_data['createdAt']:
                        task_data['createdAt'] = task_data['createdAt'].isoformat()
                    if 'updatedAt' in task_data and task_data['updatedAt']:
                        task_data['updatedAt'] = task_data['updatedAt'].isoformat()
                    
                    task_list.append(task_data)
                
                project_data['tasks'] = task_list
                project_data['is_completed'] = is_complete
                filtered_projects.append(project_data)
            else:
                print(f"Project {project_data.get('proj_name', 'Unknown')} excluded - user is not collaborator")
        
        print(f"Returning {len(filtered_projects)} filtered projects for user {current_user_id}")
        return jsonify(filtered_projects), 200
        
    except Exception as e:
        print(f"Error fetching filtered projects: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== GET FILTERED USERS BY DIVISION ===============
@projects_bp.route('/api/users/filtered/<division_name>', methods=['GET'])
def get_filtered_users_by_division(division_name):
    """Get users filtered by division"""
    try:
        db = get_firestore_client()
        
        print(f"Filtering users for division: {division_name}")
        
        users_ref = db.collection('Users')
        users_query = users_ref.where('division_name', '==', division_name)
        users = users_query.stream()
        
        user_list = []
        for user in users:
            user_data = user.to_dict()
            user_data['id'] = user.id
            
            # Convert timestamps to ISO format for JSON serialization
            if 'createdAt' in user_data and user_data['createdAt']:
                user_data['createdAt'] = user_data['createdAt'].isoformat()
            if 'updatedAt' in user_data and user_data['updatedAt']:
                user_data['updatedAt'] = user_data['updatedAt'].isoformat()
            
            user_list.append(user_data)
        
        print(f"Returning {len(user_list)} filtered users for {division_name}")
        return jsonify(user_list), 200
        
    except Exception as e:
        print(f"Error fetching filtered users: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== GET ALL PROJECTS (EXISTING) ===============
@projects_bp.route('/api/projects', methods=['GET'])
def get_all_projects_with_tasks():
    try:
        db = get_firestore_client()
        
        # Get all projects
        projects_ref = db.collection('Projects')
        projects = projects_ref.stream()
        
        project_list = []
        for project in projects:
            project_data = project.to_dict()
            project_data['id'] = project.id
            
            # Convert timestamps to ISO format for JSON serialization
            if 'start_date' in project_data and project_data['start_date']:
                project_data['start_date'] = project_data['start_date'].isoformat()
            if 'end_date' in project_data and project_data['end_date']:
                project_data['end_date'] = project_data['end_date'].isoformat()
            if 'createdAt' in project_data and project_data['createdAt']:
                project_data['createdAt'] = project_data['createdAt'].isoformat()
            if 'updatedAt' in project_data and project_data['updatedAt']:
                project_data['updatedAt'] = project_data['updatedAt'].isoformat()
            
            # Get tasks for this project using document ID
            project_doc_id = project.id
            tasks_ref = db.collection('Tasks')
            tasks_query = tasks_ref.where('proj_ID', '==', project_doc_id)
            tasks = tasks_query.stream()
            
            task_list = []
            for task in tasks:
                task_data = task.to_dict()
                task_data['id'] = task.id

                if task_data.get('is_deleted', False):
                    continue
                
                # Convert task timestamps
                if 'start_date' in task_data and task_data['start_date']:
                    task_data['start_date'] = task_data['start_date'].isoformat()
                if 'end_date' in task_data and task_data['end_date']:
                    task_data['end_date'] = task_data['end_date'].isoformat()
                if 'createdAt' in task_data and task_data['createdAt']:
                    task_data['createdAt'] = task_data['createdAt'].isoformat()
                if 'updatedAt' in task_data and task_data['updatedAt']:
                    task_data['updatedAt'] = task_data['updatedAt'].isoformat()
                
                task_list.append(task_data)
            
            project_data['tasks'] = task_list
            project_list.append(project_data)
            
        return jsonify(project_list), 200
    except Exception as e:
        print(f"Error fetching projects: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== GET SINGLE PROJECT BY ID ===============
@projects_bp.route('/api/projects/<project_id>', methods=['GET'])
def get_project_with_tasks(project_id):
    try:
        db = get_firestore_client()
        
        # Get project by document ID
        doc_ref = db.collection('Projects').document(project_id)
        doc = doc_ref.get()

        if not doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        project_data = doc.to_dict()
        project_data['id'] = doc.id
        
        # Convert timestamps to ISO format
        if 'start_date' in project_data and project_data['start_date']:
            project_data['start_date'] = project_data['start_date'].isoformat()
        if 'end_date' in project_data and project_data['end_date']:
            project_data['end_date'] = project_data['end_date'].isoformat()
        if 'createdAt' in project_data and project_data['createdAt']:
            project_data['createdAt'] = project_data['createdAt'].isoformat()
        if 'updatedAt' in project_data and project_data['updatedAt']:
            project_data['updatedAt'] = project_data['updatedAt'].isoformat()

        # Get tasks for this project using the document ID
        tasks_ref = db.collection('Tasks')
        tasks_query = tasks_ref.where('proj_ID', '==', project_id)
        tasks = tasks_query.stream()
        
        task_list = []
        for task in tasks:
            task_data = task.to_dict()
            task_data['id'] = task.id

            if task_data.get('is_deleted', False):
                continue
            
            # Convert task timestamps
            if 'start_date' in task_data and task_data['start_date']:
                task_data['start_date'] = task_data['start_date'].isoformat()
            if 'end_date' in task_data and task_data['end_date']:
                task_data['end_date'] = task_data['end_date'].isoformat()
            if 'createdAt' in task_data and task_data['createdAt']:
                task_data['createdAt'] = task_data['createdAt'].isoformat()
            if 'updatedAt' in task_data and task_data['updatedAt']:
                task_data['updatedAt'] = task_data['updatedAt'].isoformat()
            
            task_list.append(task_data)

        project_data['tasks'] = task_list
        return jsonify(project_data), 200

    except Exception as e:
        print(f"Error fetching project: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
# =============== GET ALL PROJECT BY ID'S COLLABORATORS TASKS SCHEDULE  - VIEW TEAM MEMBER'S SCHEDULE & WORKLOAD ===============
@projects_bp.route('/api/projects/<project_id>/team-schedule', methods=['GET'])
def get_project_team_schedule(project_id):
    try:
        db = get_firestore_client()
        
        # Get project details
        project_ref = db.collection('Projects').document(project_id)
        project_doc = project_ref.get()

        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        collaborator_ids = project_data.get('collaborators', [])
        
        if not collaborator_ids:
            return jsonify({
                'project': {
                    'id': project_id,
                    'proj_name': project_data.get('proj_name', 'Unknown'),
                    'start_date': project_data.get('start_date').isoformat() if project_data.get('start_date') else None,
                    'end_date': project_data.get('end_date').isoformat() if project_data.get('end_date') else None
                },
                'collaborators': [],
                'timeline_summary': {
                    'earliest_task': None,
                    'latest_task': None,
                    'total_tasks': 0,
                    'tasks_by_status': {}
                }
            }), 200

        # Get all tasks for this project
        tasks_ref = db.collection('Tasks')
        tasks_query = tasks_ref.where('proj_ID', '==', project_id)
        all_tasks = list(tasks_query.stream())
        
        # Organize tasks by user
        tasks_by_user = {}
        earliest_date = None
        latest_date = None
        status_counts = {}
        
        for task_doc in all_tasks:
            task_data = task_doc.to_dict()

            if task_data.get('is_deleted', False):
                continue

            assigned_users = task_data.get('assigned_to', [])
            
            # Count task statuses
            status = task_data.get('task_status', 'Not Started')
            status_counts[status] = status_counts.get(status, 0) + 1
            
            # Track date range
            if task_data.get('start_date'):
                if not earliest_date or task_data['start_date'] < earliest_date:
                    earliest_date = task_data['start_date']
            if task_data.get('end_date'):
                if not latest_date or task_data['end_date'] > latest_date:
                    latest_date = task_data['end_date']
            
            # Add task to each assigned user
            for user_id in assigned_users:
                if user_id not in tasks_by_user:
                    tasks_by_user[user_id] = []
                
                # Check if task is overdue
                is_overdue = False
                if task_data.get('end_date') and status not in ['Completed', 'Cancelled']:
                    from datetime import datetime
                    import pytz
                    sg_tz = pytz.timezone('Asia/Singapore')
                    now = datetime.now(sg_tz)
                    is_overdue = task_data['end_date'] < now
                
                task_info = {
                    'task_id': task_data.get('task_ID'),
                    'firestore_id': task_doc.id,
                    'task_name': task_data.get('task_name', 'Untitled Task'),
                    'task_desc': task_data.get('task_desc', ''),
                    'start_date': task_data.get('start_date').isoformat() if task_data.get('start_date') else None,
                    'end_date': task_data.get('end_date').isoformat() if task_data.get('end_date') else None,
                    'task_status': status,
                    'priority_level': task_data.get('priority_level', 'Medium'),
                    'completion_percentage': task_data.get('completion_percentage', 0),
                    'is_overdue': is_overdue,
                    'task_owner': task_data.get('task_owner')
                }
                tasks_by_user[user_id].append(task_info)
        
        # Get user details and build collaborator list
        users_ref = db.collection('Users')
        collaborators_list = []
        
        for user_id in collaborator_ids:
            user_doc = users_ref.document(user_id).get()
            if not user_doc.exists:
                continue
                
            user_data = user_doc.to_dict()
            user_tasks = tasks_by_user.get(user_id, [])
            
            # Sort tasks by start date (earliest first)
            user_tasks.sort(key=lambda t: t['start_date'] if t['start_date'] else '9999-12-31')
            
            # Calculate task statistics for this user
            completed = sum(1 for t in user_tasks if t['task_status'] == 'Completed')
            in_progress = sum(1 for t in user_tasks if t['task_status'] == 'In Progress')
            not_started = sum(1 for t in user_tasks if t['task_status'] == 'Not Started')
            
            collaborator_info = {
                'user_id': user_id,
                'name': user_data.get('name', 'Unknown User'),
                'email': user_data.get('email', ''),
                'profile_picture': user_data.get('profile_picture', ''),
                'total_tasks': len(user_tasks),
                'completed_tasks': completed,
                'in_progress_tasks': in_progress,
                'not_started_tasks': not_started,
                'overdue_tasks': sum(1 for t in user_tasks if t['is_overdue']),
                'tasks': user_tasks
            }
            collaborators_list.append(collaborator_info)
        
        # Sort collaborators by total tasks (busiest first) or alphabetically
        collaborators_list.sort(key=lambda c: (-c['total_tasks'], c['name']))
        
        # Build response
        response = {
            'project': {
                'id': project_id,
                'proj_name': project_data.get('proj_name', 'Unknown'),
                'proj_desc': project_data.get('proj_desc', ''),
                'start_date': project_data.get('start_date').isoformat() if project_data.get('start_date') else None,
                'end_date': project_data.get('end_date').isoformat() if project_data.get('end_date') else None,
                'proj_status': project_data.get('proj_status', 'Active')
            },
            'collaborators': collaborators_list,
            'timeline_summary': {
                'earliest_task': earliest_date.isoformat() if earliest_date else None,
                'latest_task': latest_date.isoformat() if latest_date else None,
                'total_tasks': len(all_tasks),
                'tasks_by_status': status_counts,
                'total_collaborators': len(collaborators_list)
            }
        }
        
        return jsonify(response), 200

    except Exception as e:
        print(f"Error fetching team schedule: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
# =============== GET PROJECT WITH TASKS (EXISTING) ===============
@projects_bp.route('/api/projects/<project_id>/tasks', methods=['GET'])
def get_project_tasks(project_id):
    try:
        db = get_firestore_client()
                
        # Verify the project exists
        project_ref = db.collection('Projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404

        project_data = project_doc.to_dict()
        project_data['id'] = project_doc.id

        # Convert timestamps in project
        if 'start_date' in project_data and project_data['start_date']:
            project_data['start_date'] = project_data['start_date'].isoformat()
        if 'end_date' in project_data and project_data['end_date']:
            project_data['end_date'] = project_data['end_date'].isoformat()
        if 'createdAt' in project_data and project_data['createdAt']:
            project_data['createdAt'] = project_data['createdAt'].isoformat()
        if 'updatedAt' in project_data and project_data['updatedAt']:
            project_data['updatedAt'] = project_data['updatedAt'].isoformat()

        # Query tasks for this project using document ID
        tasks_ref = db.collection('Tasks')
        tasks_query = tasks_ref.where('proj_ID', '==', project_id).stream()
        
        task_list = []
        task_count = 0
        
        for task in tasks_query:
            task_count += 1
            task_data = task.to_dict()
            task_data['id'] = task.id

            if task_data.get('is_deleted', False):
                continue
            
            # Convert timestamps in tasks
            if 'start_date' in task_data and task_data['start_date']:
                task_data['start_date'] = task_data['start_date'].isoformat()
            if 'end_date' in task_data and task_data['end_date']:
                task_data['end_date'] = task_data['end_date'].isoformat()
            if 'createdAt' in task_data and task_data['createdAt']:
                task_data['createdAt'] = task_data['createdAt'].isoformat()
            if 'updatedAt' in task_data and task_data['updatedAt']:
                task_data['updatedAt'] = task_data['updatedAt'].isoformat()

            task_list.append(task_data)
        
        # If no tasks found, let's check if there are ANY tasks in the Tasks collection
        if len(task_list) == 0:
            all_tasks_query = tasks_ref.limit(5).stream()  # Just get first 5 for debugging
            for task in all_tasks_query:
                task_data = task.to_dict()

        # Return project with tasks
        response = {
            **project_data,
            "tasks": task_list
        }

        return jsonify(response), 200
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== GET SPECIFIC TASK IN PROJECT (EXISTING) ===============
@projects_bp.route('/api/projects/<project_id>/tasks/<task_id>', methods=['GET'])
def get_specific_project_task(project_id, task_id):
    try:
        db = get_firestore_client()
        
        # Get the project first
        project_ref = db.collection('Projects').document(project_id)
        project_doc = project_ref.get()
        
        if not project_doc.exists:
            return jsonify({'error': 'Project not found'}), 404
            
        project_data = project_doc.to_dict()
        project_data['id'] = project_doc.id
        
        # Convert project timestamps
        if 'start_date' in project_data and project_data['start_date']:
            project_data['start_date'] = project_data['start_date'].isoformat()
        if 'end_date' in project_data and project_data['end_date']:
            project_data['end_date'] = project_data['end_date'].isoformat()
        if 'createdAt' in project_data and project_data['createdAt']:
            project_data['createdAt'] = project_data['createdAt'].isoformat()
        if 'updatedAt' in project_data and project_data['updatedAt']:
            project_data['updatedAt'] = project_data['updatedAt'].isoformat()

        # Get the specific task by document ID
        task_doc_ref = db.collection('Tasks').document(task_id)
        task_doc = task_doc_ref.get()
        
        if not task_doc.exists:
            return jsonify({'error': 'Task not found'}), 404
            
        task_data = task_doc.to_dict()
        task_data['id'] = task_doc.id
        
        # Verify the task belongs to this project
        if task_data.get('proj_ID') != project_id:
            return jsonify({'error': 'Task does not belong to this project'}), 404
        
        # Convert task timestamps
        if 'start_date' in task_data and task_data['start_date']:
            task_data['start_date'] = task_data['start_date'].isoformat()
        if 'end_date' in task_data and task_data['end_date']:
            task_data['end_date'] = task_data['end_date'].isoformat()
        if 'createdAt' in task_data and task_data['createdAt']:
            task_data['createdAt'] = task_data['createdAt'].isoformat()
        if 'updatedAt' in task_data and task_data['updatedAt']:
            task_data['updatedAt'] = task_data['updatedAt'].isoformat()
        
        # Return both project and task data
        return jsonify({
            'project': project_data,
            'task': task_data
        }), 200

    except Exception as e:
        print(f"Error fetching project task: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== GET USERS (EXISTING) ===============
@projects_bp.route('/api/users', methods=['GET'])
def get_all_users():
    try:
        db = get_firestore_client()
        users_ref = db.collection('Users')
        users = users_ref.stream()
        
        user_list = []
        for user in users:
            user_data = user.to_dict()
            user_data['id'] = user.id
            
            # Convert timestamps to ISO format for JSON serialization
            if 'createdAt' in user_data and user_data['createdAt']:
                user_data['createdAt'] = user_data['createdAt'].isoformat()
            if 'updatedAt' in user_data and user_data['updatedAt']:
                user_data['updatedAt'] = user_data['updatedAt'].isoformat()
            
            user_list.append(user_data)
            
        return jsonify(user_list), 200
    except Exception as e:
        print(f"Error fetching users: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== CREATE PROJECT ===============
@projects_bp.route('/api/projects', methods=['POST'])
def create_project():
    """Create a new project"""
    try:
        db = get_firestore_client()
        
        # Get request data
        project_data = request.get_json()
        
        if not project_data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate required fields
        required_fields = ['proj_name', 'start_date', 'end_date', 'owner', 'division_name']
        for field in required_fields:
            if field not in project_data or not project_data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Validate project name
        if len(project_data['proj_name'].strip()) < 3:
            return jsonify({'error': 'Project name must be at least 3 characters'}), 400
        
        # Validate dates
        try:
            start_date = datetime.fromisoformat(project_data['start_date'].replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(project_data['end_date'].replace('Z', '+00:00'))
            
            # Check if start date is not in the past (allow today)
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            if start_date.replace(tzinfo=None) < today:
                return jsonify({'error': 'Start date cannot be in the past'}), 400
            
            # Check if end date is after start date
            if end_date <= start_date:
                return jsonify({'error': 'End date must be after start date'}), 400
                
        except ValueError as e:
            return jsonify({'error': f'Invalid date format: {str(e)}'}), 400
        
        # Get creator ID
        owner_id = project_data['owner']
        
        # Ensure collaborators includes the creator
        collaborators = project_data.get('collaborators', [])
        if owner_id not in collaborators:
            collaborators.append(owner_id)
        
        # Prepare project data for Firestore
        firestore_data = {
            'proj_name': project_data['proj_name'].strip(),
            'proj_desc': project_data.get('proj_desc', '').strip(),
            'start_date': start_date,
            'end_date': end_date,
            'proj_status': 'Not Started',  # Default status for new projects
            'owner': owner_id,
            'division_name': project_data['division_name'],
            'collaborators': collaborators,  # 👈 INCLUDES CREATOR
            'createdAt': firestore.SERVER_TIMESTAMP,
            'updatedAt': firestore.SERVER_TIMESTAMP
        }
        
        print(f"Creating project: {firestore_data['proj_name']} with collaborators: {collaborators}")
        
        # Add project to Firestore
        doc_ref = db.collection('Projects').add(firestore_data)
        project_id = doc_ref[1].id
        
        # Get the created project back with the ID
        created_project = doc_ref[1].get().to_dict()
        created_project['id'] = project_id
        
        # Convert timestamps back to ISO format for response
        if 'start_date' in created_project and created_project['start_date']:
            created_project['start_date'] = created_project['start_date'].isoformat()
        if 'end_date' in created_project and created_project['end_date']:
            created_project['end_date'] = created_project['end_date'].isoformat()
        if 'createdAt' in created_project and created_project['createdAt']:
            created_project['createdAt'] = created_project['createdAt'].isoformat()
        if 'updatedAt' in created_project and created_project['updatedAt']:
            created_project['updatedAt'] = created_project['updatedAt'].isoformat()
        
        print(f"Project created successfully with ID: {project_id}")

        # ================== SEND EMAILS TO COLLABORATORS ==================
        try:
            from services.email_service import email_service  # import the singleton instance

            # Get creator's info (for the "Created by" field)
            creator_doc = db.collection('Users').document(owner_id).get()
            creator_name = creator_doc.to_dict().get('name', 'Unknown User') if creator_doc.exists else 'Unknown User'

            # Send emails to each collaborator (except creator)
            for collab_id in collaborators:
                if collab_id == owner_id:
                    continue

                user_doc = db.collection('Users').document(collab_id).get()
                if user_doc.exists:
                    user_data = user_doc.to_dict()
                    to_email = user_data.get('email')
                    user_name = user_data.get('name', 'User')

                    if to_email:
                        email_service.send_project_assignment_email(
                            to_email=to_email,
                            user_name=user_name,
                            project_name=firestore_data['proj_name'],
                            project_desc=firestore_data.get('proj_desc', ''),
                            creator_name=creator_name,
                            start_date=str(firestore_data['start_date'].date()),
                            end_date=str(firestore_data['end_date'].date())
                        )
                    else:
                        print(f"⚠️ No email found for user {collab_id}")
        except Exception as e:
            print(f"❌ Failed to send email notifications: {e}")
        # ================================================================
        
        return jsonify({
            'message': 'Project created successfully',
            'project': created_project
        }), 201
        
    except Exception as e:
        print(f"Error creating project: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== EXPORT PROJECT TASKS TO PDF ===============
@projects_bp.route("/api/projects/<project_id>/export", methods=["GET"])
def export_project_tasks(project_id):
    try:
        db = get_firestore_client()

        # --- Fetch Project Info ---
        project_doc = db.collection("Projects").document(project_id).get()
        project_name = "Unnamed Project"
        if project_doc.exists:
            project_data = project_doc.to_dict()
            project_name = project_data.get("proj_name", project_name)

        # --- Fetch Tasks for Project ---
        tasks_ref = db.collection("Tasks").where("proj_ID", "==", project_id)
        # tasks_ref = db.collection("Tasks").where("proj_ID", "==", project_id).order_by("end_date")
        # tasks_query = (
        #     tasks_ref.where("proj_ID", "==", project_id)
        #     # .order_by("end_date")
        # )

        tasks = [doc.to_dict() for doc in tasks_ref.stream()]
        # tasks = [doc.to_dict() for doc in tasks_query.stream()]
        tasks.sort(key=lambda task: task.get("end_date"))

        # print('HELP LAAAAA', tasks)

        if not tasks:
            return jsonify({"error": f"No tasks found for project ID: {project_id}"}), 404

        # --- Fetch all users (for name and department lookup) ---
        users_ref = db.collection("Users")
        users = {u.id: u.to_dict() for u in users_ref.stream()}

        def get_user_info(uid):
            user = users.get(uid)
            if not user:
                return "Unknown"
            name = user.get("name", "Unknown")
            dept = user.get("division_name", "N/A")
            return f"{name} ({dept})"

        # --- Compute summary ---
        total_tasks = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.get("task_status") == "Completed")
        incomplete_tasks = total_tasks - completed_tasks
        avg_priority = round(mean([t.get("priority_level", 0) for t in tasks if isinstance(t.get("priority_level"), (int, float))]), 1)

        # Get format type from request parameters
        format_type = request.args.get('format', 'table')  # 'table' or 'calendar'
        
        # --- Create PDF ---
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=40,
            rightMargin=40,
            topMargin=40,
            bottomMargin=40,
        )
        styles = getSampleStyleSheet()
        elements = []

        # --- Calendar Format ---
        if format_type == 'calendar':
            # Title
            elements.append(Paragraph(f"<b>Project Team Calendar: {project_name}</b>", styles["h1"]))
            elements.append(Spacer(1, 24))
            
            # Generate team calendar similar to Excel format
            team_calendar = generate_team_calendar_pdf(tasks, project_name, project_data, users_ref)
            for calendar_element in team_calendar:
                if isinstance(calendar_element, Table):
                    elements.append(KeepTogether(calendar_element))
                else:
                    elements.append(calendar_element)
        else:
            # --- Title ---
            elements.append(Paragraph(f"<b>Project Report: {project_name}</b>", styles["h1"]))
            # elements.append(Spacer(1, 6))
            elements.append(Paragraph(f"ID: {project_id}", styles["h3"]))
            elements.append(Spacer(1, 24))

            # --- Summary Table ---
            summary_data = [
                ["Total Tasks", "Completed", "Incomplete", "Average Priority"],
                [total_tasks, completed_tasks, incomplete_tasks, avg_priority],
            ]
            summary_table = Table(summary_data, repeatRows=1)
            summary_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightsteelblue),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            # --- Tasks Table ---
            task_data = [["Name", "Due Date", "Status", "Priority", "Owner", "Collaborators"]]

            for task in tasks:
                # name
                task_name = Paragraph(task.get("task_name", "Unknown Task"), styles['BodyText'])

                # Format date (dd-mm-yyyy)
                due_date = task.get("end_date", "—")
                if due_date and isinstance(due_date, str) and "T" not in due_date:
                    try:
                        dt = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
                        due_date = dt.strftime("%d-%m-%Y")
                    except Exception:
                        due_date = str(due_date)[:10]
                elif isinstance(due_date, datetime):
                    due_date = due_date.strftime("%d-%m-%Y")

                # Owner & collaborators
                owner_name = get_user_info(task.get("owner"))
                collaborators_list = task.get("assigned_to", [])
                collaborators_names = [get_user_info(uid) for uid in collaborators_list]
                # collaborators_str = ", ".join(collaborators_names) if collaborators_names else "—"
                collaborators_str = []
                for name in collaborators_names:
                    collaborators_str.append(Paragraph(name, styles['Normal'], bulletText='•'))

                task_data.append([
                    task_name,
                    due_date,
                    task.get("task_status", "N/A"),
                    task.get("priority_level", "—"),
                    owner_name,
                    collaborators_str,
                ])

            task_table = Table(task_data, repeatRows=1, colWidths=[90, 70, 90, 50, 100, 160])
            task_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightcoral),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]))
            elements.append(task_table)

        # --- Build PDF ---
        doc.build(elements)
        buffer.seek(0)
        
        if format_type == 'calendar':
            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"{project_name.replace(' ', '_')}_Calendar.pdf",
                mimetype="application/pdf"
            )
        else:
            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"Project_{project_name.replace(' ', '_')}_Tasks_Report.pdf",
                mimetype="application/pdf"
            )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =============== EXPORT PROJECT TEAM SCHEDULE TO EXCEL ===============
@projects_bp.route("/api/projects/<project_id>/export-excel", methods=["GET"])
def export_project_team_schedule_excel(project_id):
    try:
        print(f"=== EXCEL EXPORT DEBUG ===")
        print(f"Project ID: {project_id}")
        
        db = get_firestore_client()
        print("Firestore client obtained")

        # --- Fetch Project Info ---
        project_doc = db.collection("Projects").document(project_id).get()
        project_name = "Unnamed Project"
        if project_doc.exists:
            project_data = project_doc.to_dict()
            project_name = project_data.get("proj_name", project_name)
            print(f"Project found: {project_name}")
        else:
            print(f"Project not found for ID: {project_id}")

        # --- Fetch Tasks for Project ---
        tasks_ref = db.collection("Tasks").where("proj_ID", "==", project_id)
        tasks = [doc.to_dict() for doc in tasks_ref.stream()]
        tasks.sort(key=lambda task: task.get("end_date"))
        print(f"Found {len(tasks)} tasks")

        if not tasks:
            print("No tasks found, returning 404")
            return jsonify({"error": f"No tasks found for project ID: {project_id}"}), 404

        # --- Fetch all users (for name and department lookup) ---
        users_ref = db.collection("Users")
        users = {u.id: u.to_dict() for u in users_ref.stream()}

        def get_user_info(uid):
            user = users.get(uid)
            if not user:
                return "Unknown"
            name = user.get("name", "Unknown")
            dept = user.get("division_name", "N/A")
            return f"{name} ({dept})"

        # --- Create Excel Workbook ---
        print("Creating Excel workbook...")
        wb = Workbook()
        # Remove the default sheet - we'll create monthly sheets directly
        wb.remove(wb.active)
        print("Excel workbook created")

        # --- Define Premium Styles ---
        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Deep blue
        month_header_font = Font(bold=True, color="1E3A8A", size=18, name="Arial")
        
        # Border styles
        thick_border = Border(
            left=Side(style='thick', color="1E3A8A"),
            right=Side(style='thick', color="1E3A8A"),
            top=Side(style='thick', color="1E3A8A"),
            bottom=Side(style='thick', color="1E3A8A")
        )
        thin_border = Border(
            left=Side(style='thin', color="E5E7EB"),
            right=Side(style='thin', color="E5E7EB"),
            top=Side(style='thin', color="E5E7EB"),
            bottom=Side(style='thin', color="E5E7EB")
        )
        
        # Alignment styles
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Premium priority colors with gradients
        high_priority_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")  # Red
        medium_priority_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")  # Amber
        low_priority_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")  # Emerald
        completed_fill = PatternFill(start_color="6B7280", end_color="6B7280", fill_type="solid")  # Gray
        
        # Collaborator colors - each team member gets a unique color
        collaborator_colors = [
            "FF6B6B",  # Red
            "4ECDC4",  # Teal
            "45B7D1",  # Blue
            "96CEB4",  # Green
            "FFEAA7",  # Yellow
            "DDA0DD",  # Plum
            "98D8C8",  # Mint
            "F7DC6F",  # Light Yellow
            "BB8FCE",  # Light Purple
            "85C1E9",  # Light Blue
            "F8C471",  # Light Orange
            "82E0AA",  # Light Green
            "F1948A",  # Light Red
            "85C1E9",  # Sky Blue
            "D7BDE2"   # Light Lavender
        ]
        
        # Background colors
        member_cell_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # Light gray
        weekend_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Light yellow
        today_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Light blue


        # --- Create Monthly Sheets ---
        print("Creating monthly sheets...")
        
        # Import calendar module
        import calendar as cal_module
        
        # Get all unique dates from tasks and group by month
        dates_by_month = {}
        for task in tasks:
            start_date = task.get("start_date", "")
            end_date = task.get("end_date", "")
            
            print(f"Task: {task.get('task_name', '')}")
            print(f"  Start date: {start_date} (type: {type(start_date)})")
            print(f"  End date: {end_date} (type: {type(end_date)})")
            
            if start_date:
                try:
                    if isinstance(start_date, str):
                        if "T" in start_date:
                            dt = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                        else:
                            dt = datetime.fromisoformat(start_date)
                    else:
                        dt = start_date
                    
                    if dt.tzinfo:
                        dt = dt.replace(tzinfo=None)
                    
                    month_key = (dt.year, dt.month)
                    if month_key not in dates_by_month:
                        dates_by_month[month_key] = set()
                    dates_by_month[month_key].add(dt.date())
                    print(f"  Parsed start date: {dt.date()}")
                except Exception as e:
                    print(f"  Error parsing start date: {e}")
            
            if end_date:
                try:
                    if isinstance(end_date, str):
                        if "T" in end_date:
                            dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                        else:
                            dt = datetime.fromisoformat(end_date)
                    else:
                        dt = end_date
                    
                    if dt.tzinfo:
                        dt = dt.replace(tzinfo=None)
                    
                    month_key = (dt.year, dt.month)
                    if month_key not in dates_by_month:
                        dates_by_month[month_key] = set()
                    dates_by_month[month_key].add(dt.date())
                    print(f"  Parsed end date: {dt.date()}")
                except Exception as e:
                    print(f"  Error parsing end date: {e}")
        
        print(f"Found {len(dates_by_month)} months with tasks")
        
        if not dates_by_month:
            print("No dates found, creating empty sheet")
            # Create a single sheet for when there are no tasks
            ws = wb.create_sheet("No Tasks")
            ws['A1'] = f"📊 Project: {project_name}"
            ws['A1'].font = Font(bold=True, size=16, name="Arial", color="1E3A8A")
            ws['A8'] = "No tasks with dates found"
            ws['A8'].font = Font(italic=True, color="666666")
        else:
            
            # Create a separate sheet for each month
            for (year, month), month_dates in sorted(dates_by_month.items()):
                print(f"Creating sheet for {cal_module.month_name[month]} {year}")
                
                # Create new worksheet for this month
                month_sheet_name = f"{cal_module.month_name[month][:3]} {year}"
                if month_sheet_name in wb.sheetnames:
                    ws = wb[month_sheet_name]
                else:
                    ws = wb.create_sheet(title=month_sheet_name)
                
                # Add premium header to this sheet
                ws['A1'] = f"📊 {cal_module.month_name[month]} {year} - {project_name}"
                ws['A1'].font = Font(bold=True, size=16, name="Arial", color="1E3A8A")
                ws['A1'].fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
                ws['A1'].border = thick_border
                ws['A1'].alignment = center_alignment
                
                ws['A2'] = f"🕒 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A2'].font = Font(size=11, italic=True, name="Arial", color="6B7280")
                ws['A2'].alignment = center_alignment
                
                # Add legend header
                ws['A3'] = "👥 Team Members:"
                ws['A3'].font = Font(bold=True, size=12, name="Arial", color="1E3A8A")
                
                # Calculate month days
                month_days = cal_module.monthrange(year, month)[1]
                
                # First, collect all tasks for this month
                month_tasks = []
                for task in tasks:
                    assigned_to = task.get("assigned_to", [])
                    start_date = task.get("start_date", "")
                    end_date = task.get("end_date", "")
                    
                    task_start = None
                    task_end = None
                    
                    if start_date:
                        try:
                            if isinstance(start_date, str):
                                if "T" in start_date:
                                    dt = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
                                else:
                                    dt = datetime.fromisoformat(start_date)
                            else:
                                dt = start_date
                            
                            if dt.tzinfo:
                                dt = dt.replace(tzinfo=None)
                            task_start = dt.date()
                        except:
                            pass
                    
                    if end_date:
                        try:
                            if isinstance(end_date, str):
                                if "T" in end_date:
                                    dt = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                                else:
                                    dt = datetime.fromisoformat(end_date)
                            else:
                                dt = end_date
                            
                            if dt.tzinfo:
                                dt = dt.replace(tzinfo=None)
                            task_end = dt.date()
                        except:
                            pass
                    
                    # Check if task overlaps with this month
                    if task_start and task_end:
                        month_start = datetime(year, month, 1).date()
                        month_end = datetime(year, month, cal_module.monthrange(year, month)[1]).date()
                        
                        # Task overlaps with this month if it starts before month ends and ends after month starts
                        if task_start <= month_end and task_end >= month_start:
                            # Get team members for this task
                            for user_id in assigned_to:
                                # Get user name from Firestore
                                try:
                                    user_doc = users_ref.document(user_id).get()
                                    if user_doc.exists:
                                        member_name = user_doc.to_dict().get('name', f'User_{user_id}')
                                    else:
                                        member_name = f'User_{user_id}'
                                except:
                                    member_name = f'User_{user_id}'
                                
                                month_tasks.append({
                                    'name': task.get('task_name', ''),
                                    'member': member_name,
                                    'member_id': user_id,  # Store user ID for filtering
                                    'start': task_start,
                                    'end': task_end,
                                    'priority': task.get('priority_level', 1)  # Use consistent field name
                                })
                
                # Sort tasks by start date
                month_tasks.sort(key=lambda x: x['start'])
                print(f"Found {len(month_tasks)} tasks for {cal_module.month_name[month]} {year}")
                
                # Create Gantt chart with date headers
                current_row = 4  # Start after project header
                
                # Add priority color legend first
                ws[f'A{current_row}'] = "🎯 Priority Legend:"
                ws[f'A{current_row}'].font = Font(bold=True, size=12, name="Arial", color="1E3A8A")
                current_row += 1
                
                # High priority
                ws[f'A{current_row}'] = "🔴 High Priority (4-5)"
                ws[f'A{current_row}'].font = Font(size=10, name="Arial")
                ws[f'A{current_row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = center_alignment
                current_row += 1
                
                # Medium priority
                ws[f'A{current_row}'] = "🟡 Medium Priority (2-3)"
                ws[f'A{current_row}'].font = Font(size=10, name="Arial")
                ws[f'A{current_row}'].fill = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = center_alignment
                current_row += 1
                
                # Low priority
                ws[f'A{current_row}'] = "🟢 Low Priority (1)"
                ws[f'A{current_row}'].font = Font(size=10, name="Arial")
                ws[f'A{current_row}'].fill = PatternFill(start_color="96CEB4", end_color="96CEB4", fill_type="solid")
                ws[f'A{current_row}'].border = thin_border
                ws[f'A{current_row}'].alignment = center_alignment
                current_row += 1
                
                ws[f'A{current_row}'] = ""  # Empty row after legend
                current_row += 1
                
                # Create day-of-week headers starting from column B
                day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
                for i, day_name in enumerate(day_names):
                    col = i + 2  # Start from column B (column 2)
                    day_cell = ws.cell(row=current_row, column=col, value=day_name)
                    day_cell.font = header_font
                    day_cell.fill = header_fill
                    day_cell.border = thick_border
                    day_cell.alignment = center_alignment
                current_row += 1
                
                # Group tasks by team member and filter by project collaborators
                tasks_by_member = {}
                
                # Get project collaborators from project data
                project_collaborator_ids = set(project_data.get('collaborators', []))
                print(f"Project collaborators: {project_collaborator_ids}")
                
                # Only include tasks from project collaborators (compare by user ID)
                for task in month_tasks:
                    # Check if this task's member is a project collaborator
                    if task['member_id'] in project_collaborator_ids:
                        member = task['member']
                        if member not in tasks_by_member:
                            tasks_by_member[member] = []
                        tasks_by_member[member].append(task)
                        print(f"Included task '{task['name']}' for collaborator '{member}' (ID: {task['member_id']})")
                    else:
                        print(f"Excluded task '{task['name']}' for non-collaborator '{member}' (ID: {task['member_id']})")
                
                
                # Create calendar grid with Gantt bars
                cal = cal_module.monthcalendar(year, month)
                
                # Create calendar rows (weeks) with day numbers and task bars
                for week_idx, week in enumerate(cal):
                    # Create day number row
                    for day_idx, day_num in enumerate(week):
                        col = day_idx + 2  # Start from column B (column 2)
                        
                        if day_num == 0:  # Empty day (not in this month)
                            # Empty cell
                            empty_cell = ws.cell(row=current_row, column=col, value="")
                            empty_cell.border = thin_border
                            empty_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        else:
                            # Day number cell
                            day_cell = ws.cell(row=current_row, column=col, value=day_num)
                            day_cell.font = Font(bold=True, size=12, name="Arial")
                            day_cell.border = thin_border
                            day_cell.alignment = center_alignment
                            
                            # Highlight weekends
                            date_obj = datetime(year, month, day_num)
                            if date_obj.weekday() >= 5:  # Saturday = 5, Sunday = 6
                                day_cell.fill = weekend_fill
                                day_cell.font = Font(bold=True, color="1E3A8A", size=12, name="Arial")
                            else:
                                day_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    current_row += 1
                    
                    # Create task bars row for this week
                    # Group tasks by member for this week
                    week_tasks_by_member = {}
                    for task in month_tasks:
                        task_start_day = task['start'].day if task['start'].month == month else 1
                        task_end_day = task['end'].day if task['end'].month == month else month_days
                        
                        # Check if task overlaps with this week
                        week_days = [day for day in week if day != 0]
                        if week_days and task_start_day <= max(week_days) and task_end_day >= min(week_days):
                            member = task['member']
                            if member not in week_tasks_by_member:
                                week_tasks_by_member[member] = []
                            week_tasks_by_member[member].append(task)
                    
                    # Create task bars for each member in this week
                    for member, member_tasks in week_tasks_by_member.items():
                        # Member name in first column
                        member_cell = ws.cell(row=current_row, column=1, value=member)
                        member_cell.font = Font(bold=True, size=10, name="Arial", color="1E3A8A")
                        member_cell.border = thin_border
                        member_cell.fill = member_cell_fill
                        member_cell.alignment = center_alignment
                        
                        # Track which columns are already used for this row
                        used_columns = set()
                        
                        # Create task bars across the week
                        for task in member_tasks:
                            task_start_day = task['start'].day if task['start'].month == month else 1
                            task_end_day = task['end'].day if task['end'].month == month else month_days
                            
                            # Find start and end columns within this week
                            start_col = None
                            end_col = None
                            
                            for day_idx, day_num in enumerate(week):
                                if day_num != 0:  # Only consider days in this month
                                    if day_num == task_start_day:
                                        start_col = day_idx + 2  # Start from column B (column 2)
                                    if day_num == task_end_day:
                                        end_col = day_idx + 2  # Start from column B (column 2)
                            
                            # If task starts before this week, start from first day
                            if start_col is None and task_start_day < min([d for d in week if d != 0]):
                                start_col = 2  # Column B
                            
                            # If task ends after this week, end at last day
                            if end_col is None and task_end_day > max([d for d in week if d != 0]):
                                end_col = 8  # Column H (last day column)
                            
                            if start_col and end_col:
                                # Check for overlaps
                                task_columns = set(range(start_col, end_col + 1))
                                if task_columns.intersection(used_columns):
                                    continue
                                
                                # Mark columns as used
                                used_columns.update(task_columns)
                                
                                # Create task bar with priority-based color
                                priority = task.get('priority', 1)  # This should match the field name used above
                                
                                # Convert priority to integer if it's a string
                                try:
                                    priority = int(priority)
                                except (ValueError, TypeError):
                                    priority = 1  # Default to low priority if conversion fails
                                
                                print(f"Task '{task['name']}' has priority: {priority} (type: {type(priority)})")
                                
                                # Handle different priority scales (1-3 or 1-5)
                                if priority >= 4:  # High priority (4-5 scale or 3 scale)
                                    task_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                                    print(f"  -> Using HIGH priority color (red)")
                                elif priority >= 2:  # Medium priority (2-3 scale or 2 scale)
                                    task_fill = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")
                                    print(f"  -> Using MEDIUM priority color (yellow)")
                                else:  # Low priority (1 or default)
                                    task_fill = PatternFill(start_color="96CEB4", end_color="96CEB4", fill_type="solid")
                                    print(f"  -> Using LOW priority color (green)")
                                
                                if start_col == end_col:
                                    # Single day task
                                    cell = ws.cell(row=current_row, column=start_col, value=task['name'])
                                    cell.border = thin_border
                                    cell.alignment = center_alignment
                                    cell.fill = task_fill
                                    cell.font = Font(size=8, bold=True, color="FFFFFF", name="Arial")
                                else:
                                    # Multi-day task - create horizontal bar
                                    try:
                                        start_cell = ws.cell(row=current_row, column=start_col, value=task['name'])
                                        start_cell.border = thin_border
                                        start_cell.alignment = center_alignment
                                        start_cell.fill = task_fill
                                        start_cell.font = Font(size=8, bold=True, color="FFFFFF", name="Arial")
                                        
                                        # Merge cells to create the bar
                                        ws.merge_cells(start_row=current_row, start_column=start_col, 
                                                     end_row=current_row, end_column=end_col)
                                        print(f"Successfully merged cells for task '{task['name']}'")
                                    except Exception as e:
                                        print(f"Error creating task bar for '{task['name']}': {e}")
                                        ws.cell(row=current_row, column=start_col, value=task['name'])
                        
                        current_row += 1
                
                # Calendar view complete - tasks are now displayed inside each day cell
                
                # Set column widths and row heights for this sheet
                # Calendar columns (Mon-Sun) + team member column
                ws.column_dimensions['A'].width = 20  # Team member column
                for col in range(2, 9):  # 7 days of the week (columns B-H)
                    column_letter = ws.cell(row=1, column=col).column_letter
                    ws.column_dimensions[column_letter].width = 15  # Calendar day columns
                
                # Set row heights
                for row in range(1, current_row + 5):
                    if row == 1:  # Project header
                        ws.row_dimensions[row].height = 35
                    elif row == 2:  # Timestamp
                        ws.row_dimensions[row].height = 25
                    elif row >= 8:  # Calendar rows (after headers and legend)
                        ws.row_dimensions[row].height = 40  # Height for calendar with task bars
                    else:
                        ws.row_dimensions[row].height = 25

        # --- Multi-sheet setup complete ---
        # Each month now has its own sheet with proper sizing

        # --- Save to BytesIO ---
        print("Saving Excel to buffer...")
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        print(f"Excel saved, buffer size: {buffer.getbuffer().nbytes} bytes")

        filename = f"{project_name.replace(' ', '_')}_Team_Calendar.xlsx"
        print(f"Returning file: {filename}")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"❌ EXCEL EXPORT ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    
# =============== UPDATE PROJECT ===============
@projects_bp.route('/api/projects/<project_id>', methods=['PUT'])
def update_project(project_id):
    """Update an existing project"""
    try:
        db = get_firestore_client()
        
        # Check if project exists
        doc_ref = db.collection('Projects').document(project_id)
        doc = doc_ref.get()
        
        if not doc.exists:
            return jsonify({'error': 'Project not found'}), 404
        
        existing_data = doc.to_dict()
        
        # Get request data
        update_data = request.get_json()
        
        if not update_data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Validate required fields
        required_fields = ['proj_name', 'start_date', 'end_date', 'owner', 'collaborators']
        for field in required_fields:
            if field not in update_data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Validate project name
        if not update_data['proj_name'] or len(update_data['proj_name'].strip()) < 3:
            return jsonify({'error': 'Project name must be at least 3 characters'}), 400
        
        # Validate collaborators (at least 1 required)
        if not update_data['collaborators'] or len(update_data['collaborators']) == 0:
            return jsonify({'error': 'At least 1 collaborator is required'}), 400
        
        # Validate dates
        try:
            start_date = datetime.fromisoformat(update_data['start_date'].replace('Z', '+00:00'))
            end_date = datetime.fromisoformat(update_data['end_date'].replace('Z', '+00:00'))
            
            # Check if start date is not in the past (allow today)
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            if start_date.replace(tzinfo=None) < today:
                return jsonify({'error': 'Start date cannot be in the past'}), 400
            
            # Check if end date is after start date
            if end_date <= start_date:
                return jsonify({'error': 'End date must be after start date'}), 400
                
        except ValueError as e:
            return jsonify({'error': f'Invalid date format: {str(e)}'}), 400
        
        # Ensure owner is in collaborators list
        collaborators = update_data['collaborators']
        owner_id = update_data['owner']
        if owner_id not in collaborators:
            collaborators.append(owner_id)
        
        # Prepare update data for Firestore
        firestore_update = {
            'proj_name': update_data['proj_name'].strip(),
            'proj_desc': update_data.get('proj_desc', '').strip(),
            'start_date': start_date,
            'end_date': end_date,
            'owner': owner_id,
            'division_name': update_data.get('division_name', existing_data.get('division_name', '')),
            'collaborators': collaborators,
            'updatedAt': firestore.SERVER_TIMESTAMP
        }
        
        # Update status if provided
        if 'proj_status' in update_data:
            valid_statuses = ['Not Started', 'In Progress', 'Completed', 'On Hold']
            if update_data['proj_status'] in valid_statuses:
                firestore_update['proj_status'] = update_data['proj_status']
        
        print(f"Updating project {project_id} with data: {firestore_update}")
        
        # Update the document in Firestore
        doc_ref.update(firestore_update)
        
        # Get updated project
        updated_doc = doc_ref.get()
        updated_project = updated_doc.to_dict()
        updated_project['id'] = project_id
        
        # Convert timestamps for JSON response
        if 'start_date' in updated_project and updated_project['start_date']:
            updated_project['start_date'] = updated_project['start_date'].isoformat()
        if 'end_date' in updated_project and updated_project['end_date']:
            updated_project['end_date'] = updated_project['end_date'].isoformat()
        if 'createdAt' in updated_project and updated_project['createdAt']:
            updated_project['createdAt'] = updated_project['createdAt'].isoformat()
        if 'updatedAt' in updated_project and updated_project['updatedAt']:
            updated_project['updatedAt'] = updated_project['updatedAt'].isoformat()
        
        print(f"✅ Project {project_id} updated successfully")
        
        return jsonify({
            'message': 'Project updated successfully',
            'project': updated_project
        }), 200
        
    except Exception as e:
        print(f"Error updating project: {str(e)}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# =============== EXPORT PROJECT TASKS TO EXCEL ===============
@projects_bp.route("/api/projects/<project_id>/export/xlsx", methods=["GET"])
def export_project_tasks_xlsx(project_id):
    try:
        db = get_firestore_client()

        # --- Fetch project info ---
        project_doc = db.collection("Projects").document(project_id).get()
        project_name = "Unnamed Project"
        if project_doc.exists:
            project_data = project_doc.to_dict()
            project_name = project_data.get("proj_name", project_name)

        # --- Fetch tasks ---
        tasks_ref = db.collection("Tasks").where("proj_ID", "==", project_id)
        tasks = [doc.to_dict() for doc in tasks_ref.stream()]
        tasks.sort(key=lambda task: task.get("end_date"))

        if not tasks:
            return jsonify({"error": f"No tasks found for project ID: {project_id}"}), 404

        # --- Fetch user info ---
        users_ref = db.collection("Users")
        users = {u.id: u.to_dict() for u in users_ref.stream()}

        def get_user_info(uid):
            user = users.get(uid)
            if not user:
                return "Unknown"
            name = user.get("name", "Unknown")
            dept = user.get("division_name", "N/A")
            return f"{name} ({dept})"

        # --- Summary statistics ---
        total_tasks = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.get("task_status") == "Completed")
        incomplete_tasks = total_tasks - completed_tasks
        avg_priority = round(mean([t.get("priority_level", 0) for t in tasks if isinstance(t.get("priority_level"), (int, float))]), 1)

        # --- Prepare workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Tasks"

        # --- Title section ---
        ws["A1"] = f"Project Report: {project_name}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = f"Project ID: {project_id}"
        ws["A2"].font = Font(italic=True)

        # --- Summary section ---
        ws.append([])
        summary_headers = ["Total Tasks", "Completed", "Incomplete", "Average Priority"]
        summary_values = [total_tasks, completed_tasks, incomplete_tasks, avg_priority]

        ws.append(summary_headers)
        ws.append(summary_values)

        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center")

        for cell in ws[5]:
            cell.alignment = Alignment(horizontal="center")

        ws.append([])

        # --- Task Table ---
        headers = ["Name", "Due Date", "Status", "Priority", "Owner", "Collaborators"]
        ws.append(headers)
        for cell in ws[7]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center")

        for task in tasks:
            task_name = task.get("task_name")
            due_date = task.get("end_date")
            if isinstance(due_date, str):
                try:
                    dt = datetime.fromisoformat(due_date.replace("Z", "+00:00"))
                    due_date = dt.strftime("%d-%m-%Y")
                except Exception:
                    due_date = due_date[:10]
            elif isinstance(due_date, datetime):
                due_date = due_date.strftime("%d-%m-%Y")

            owner_name = get_user_info(task.get("owner"))
            collaborators_list = task.get("assigned_to", [])
            collaborators_str = ", ".join([get_user_info(uid) for uid in collaborators_list]) if collaborators_list else "—"

            ws.append([
                task_name,
                due_date,
                task.get("task_status", "N/A"),
                task.get("priority_level", "—"),
                owner_name,
                collaborators_str,
            ])

        # --- Auto column widths ---
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # --- Add Charts ---
        ws.append([])
        ws.append(["Tasks by Status"])
        last_row_index = ws.max_row

        # for cell in ws[last_row_index]:
        ws[last_row_index][0].font = Font(bold=True)
        ws[last_row_index][0].fill = PatternFill("solid", fgColor="DDDDDD")
        ws[last_row_index][0].alignment = Alignment(horizontal="center")
        chart_start_row = ws.max_row + 2

        # Count tasks by status and priority
        from collections import Counter
        status_counts = Counter(t.get("task_status", "Unknown") for t in tasks)
        priority_counts = Counter(
            "High" if t.get("priority_level", 0) >= 7 else
            "Medium" if t.get("priority_level", 0) >= 4 else
            "Low"
            for t in tasks
        )

        # Insert status data
        ws.append(["Status", "Count"])
        for k, v in status_counts.items():
            ws.append([k, v])
        status_table_end = ws.max_row

        # Create pie chart for Status
        pie_status = PieChart()
        labels = Reference(ws, min_col=1, min_row=status_table_end - len(status_counts) + 1, max_row=status_table_end)
        data = Reference(ws, min_col=2, min_row=status_table_end - len(status_counts), max_row=status_table_end)
        pie_status.add_data(data, titles_from_data=True)
        pie_status.set_categories(labels)
        pie_status.title = "Tasks by Status"
        ws.add_chart(pie_status, f"E{chart_start_row}")

        # Insert Priority data
        ws.append([])
        ws.append(["Tasks by Priority"])
        last_row_index = ws.max_row
        ws[last_row_index][0].font = Font(bold=True)
        ws[last_row_index][0].fill = PatternFill("solid", fgColor="DDDDDD")
        ws[last_row_index][0].alignment = Alignment(horizontal="center")
        ws.append(["Priority Level", "Count"])
        for k, v in priority_counts.items():
            ws.append([k, v])
        priority_table_end = ws.max_row

        # Create pie chart for Priority
        pie_priority = PieChart()
        labels = Reference(ws, min_col=1, min_row=priority_table_end - len(priority_counts) + 1, max_row=priority_table_end)
        data = Reference(ws, min_col=2, min_row=priority_table_end - len(priority_counts), max_row=priority_table_end)
        pie_priority.add_data(data, titles_from_data=True)
        pie_priority.set_categories(labels)
        pie_priority.title = "Tasks by Priority"
        ws.add_chart(pie_priority, f"E{chart_start_row + 15}")

        # --- Save workbook to memory ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)


        return send_file(
            output,
            as_attachment=True,
            download_name=f"Project_{project_name.replace(' ', '_')}_Tasks_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500